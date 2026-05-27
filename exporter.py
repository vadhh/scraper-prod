import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

log = logging.getLogger(__name__)


def save_to_excel(books, filename="books.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    headers = ["Title", "Price", "Availability", "Star Rating", "Category", "URL"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, book in enumerate(books, 2):
        ws.cell(row=row_idx, column=1, value=book["Title"]).border = thin_border
        price_cell = ws.cell(row=row_idx, column=2, value=book["Price"])
        price_cell.number_format = "£#,##0.00"
        price_cell.border = thin_border
        ws.cell(row=row_idx, column=3, value=book["Availability"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=book["Star Rating"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=book["Category"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=book["URL"]).border = thin_border

        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    col_widths = {"A": 50, "B": 12, "C": 15, "D": 12, "E": 25, "F": 70}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.auto_filter.ref = f"A1:F{len(books) + 1}"
    ws.freeze_panes = "A2"

    wb.save(filename)
    log.info(f"Saved {len(books)} books to {filename}")
