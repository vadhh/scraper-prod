import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
import sys

BASE_URL = "https://books.toscrape.com/"
CATALOGUE_URL = BASE_URL + "catalogue/"

RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5,
}


def get_soup(url):
    response = requests.get(url, timeout=15)
    response.raise_for_status()
    return BeautifulSoup(response.text, "lxml")


def get_categories():
    soup = get_soup(BASE_URL)
    sidebar = soup.find("ul", class_="nav-list")
    category_links = sidebar.find("ul").find_all("a")
    categories = []
    for link in category_links:
        name = link.text.strip()
        href = BASE_URL + link["href"]
        categories.append((name, href))
    return categories


def parse_books_from_page(soup, category_name):
    books = []
    articles = soup.find_all("article", class_="product_pod")
    for article in articles:
        title_tag = article.find("h3").find("a")
        title = title_tag["title"]

        relative_href = title_tag["href"]
        if relative_href.startswith("../../"):
            url = CATALOGUE_URL + relative_href.replace("../", "")
        elif relative_href.startswith("../"):
            url = CATALOGUE_URL + relative_href.replace("../", "")
        else:
            url = CATALOGUE_URL + relative_href

        price_text = article.find("p", class_="price_color").text.strip()
        price = float(price_text.replace("£", "").replace("Â", ""))

        availability_tag = article.find("p", class_="instock")
        availability = "In Stock" if availability_tag else "Out of Stock"

        rating_tag = article.find("p", class_="star-rating")
        rating_class = [c for c in rating_tag["class"] if c != "star-rating"][0]
        star_rating = RATING_MAP.get(rating_class, 0)

        books.append({
            "Title": title,
            "Price": price,
            "Availability": availability,
            "Star Rating": star_rating,
            "Category": category_name,
            "URL": url,
        })
    return books


def scrape_category(name, url):
    books = []
    page_url = url
    page_num = 1

    while page_url:
        soup = get_soup(page_url)
        page_books = parse_books_from_page(soup, name)
        books.extend(page_books)
        print(f"  Page {page_num}: {len(page_books)} books")

        next_btn = soup.find("li", class_="next")
        if next_btn:
            next_href = next_btn.find("a")["href"]
            page_url = url.rsplit("/", 1)[0] + "/" + next_href
            page_num += 1
            time.sleep(0.1)
        else:
            page_url = None

    return books


def remove_duplicates(books):
    seen = set()
    unique = []
    for book in books:
        if book["URL"] not in seen:
            seen.add(book["URL"])
            unique.append(book)
    return unique


def save_to_excel(books, filename="books.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    headers = ["Title", "Price", "Availability", "Star Rating", "Category", "URL"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, book in enumerate(books, 2):
        ws.cell(row=row_idx, column=1, value=book["Title"]).border = thin_border
        price_cell = ws.cell(row=row_idx, column=2, value=book["Price"])
        price_cell.number_format = "£#,##0.00"
        price_cell.border = thin_border
        ws.cell(row=row_idx, column=3, value=book["Availability"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=book["Star Rating"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=book["Category"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=book["URL"]).border = thin_border

        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    col_widths = {"A": 50, "B": 12, "C": 15, "D": 12, "E": 25, "F": 70}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.auto_filter.ref = f"A1:F{len(books) + 1}"
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\nSaved {len(books)} books to {filename}")


def main():
    print("=" * 60)
    print("Books to Scrape - Web Scraper")
    print("=" * 60)

    print("\nFetching categories...")
    categories = get_categories()
    print(f"Found {len(categories)} categories\n")

    all_books = []
    for i, (name, url) in enumerate(categories, 1):
        print(f"[{i}/{len(categories)}] Scraping: {name}")
        books = scrape_category(name, url)
        all_books.extend(books)
        time.sleep(0.1)

    print(f"\nTotal books scraped: {len(all_books)}")

    all_books = remove_duplicates(all_books)
    print(f"After deduplication: {len(all_books)}")

    all_books.sort(key=lambda b: (b["Category"].lower(), b["Title"].lower()))

    save_to_excel(all_books)
    print("Done!")


if __name__ == "__main__":
    main()
